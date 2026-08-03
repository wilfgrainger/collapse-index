from pathlib import Path
from urllib.parse import urljoin, urlparse
import json
import requests
from bs4 import BeautifulSoup

OUT = Path("discovery")
RAW = OUT / "raw"
OUT.mkdir(exist_ok=True)
RAW.mkdir(exist_ok=True)
SESSION = requests.Session()
SESSION.headers.update({
    "User-Agent": "UK-Stability-Monitor-source-validation/0.1 (+https://github.com/wilfgrainger/collapse-index)"
})
REPORT = []

PAGES = {
    "fsa_food_catalogue": "https://data.food.gov.uk/catalog/datasets/e8563667-8457-4662-9642-aeed6454113f",
    "nhs_rtt": "https://www.england.nhs.uk/statistics/statistical-work-areas/rtt-waiting-times/rtt-data-2026-27/",
    "ons_housing_catalogue": "https://www.ons.gov.uk/peoplepopulationandcommunity/housing/datasets/privaterentalaffordabilityengland",
    "ons_trust_catalogue": "https://www.ons.gov.uk/peoplepopulationandcommunity/wellbeing/datasets/trustingovernmentuk",
    "ons_wellbeing_catalogue": "https://www.ons.gov.uk/peoplepopulationandcommunity/wellbeing/datasets/ukmeasuresofnationalwellbeing",
    "nrw_portal": "https://api-portal.naturalresources.wales/",
}
DIRECT = {
    "child_poverty": "https://www.ons.gov.uk/explore-local-statistics/api/v1/data.csv?indicator=children-in-relative-poverty-after-housing&time=all",
    "ea_floods": "https://environment.data.gov.uk/flood-monitoring/id/floods",
    "fsa_food": "https://fsaopendata.blob.core.windows.net/opendatacatalog/24-077525-01_Food%26You2Trends_W1-11_Abridged%20Data_V1.csv",
    "ons_housing_2024": "https://www.ons.gov.uk/file?uri=/peoplepopulationandcommunity/housing/datasets/privaterentalaffordabilityengland/2024/privaterentalaffordability2024.xlsx",
    "ons_housing_2023": "https://www.ons.gov.uk/file?uri=/peoplepopulationandcommunity/housing/datasets/privaterentalaffordabilityengland/2023/privaterentalaffordability2023englandwales.xlsx",
    "ons_trust_2023": "https://www.ons.gov.uk/file?uri=/peoplepopulationandcommunity/wellbeing/datasets/trustingovernmentuk/2023/bulletintrustingovernmentreftables2023.xlsx",
    "ons_trust_2022": "https://www.ons.gov.uk/file?uri=/peoplepopulationandcommunity/wellbeing/datasets/trustingovernmentuk/2022/bulletintrustingovernmentreftables.xlsx",
    "ons_wellbeing_2026": "https://www.ons.gov.uk/file?uri=/peoplepopulationandcommunity/wellbeing/datasets/ukmeasuresofnationalwellbeing/june2026/ukmeasuresofnationalwellbeingjune2026.xlsx",
}


def safe_name(name, url, content_type=""):
    suffix = Path(urlparse(url).path).suffix
    if not suffix:
        if "csv" in content_type:
            suffix = ".csv"
        elif "json" in content_type:
            suffix = ".json"
        elif "excel" in content_type or "spreadsheet" in content_type:
            suffix = ".xlsx"
        else:
            suffix = ".bin"
    return RAW / f"{name}{suffix}"


def get(url):
    response = SESSION.get(url, timeout=60, allow_redirects=True)
    REPORT.append(
        f"FETCH {url}\n  status={response.status_code} final={response.url} "
        f"type={response.headers.get('content-type')} bytes={len(response.content)}"
    )
    response.raise_for_status()
    return response


selected = {}
for name, url in DIRECT.items():
    try:
        response = get(url)
        path = safe_name(name, response.url, response.headers.get("content-type", ""))
        path.write_bytes(response.content)
        selected[name] = {
            "url": response.url,
            "path": str(path),
            "content_type": response.headers.get("content-type"),
            "bytes": len(response.content),
        }
    except Exception as error:
        REPORT.append(f"ERROR {name}: {type(error).__name__}: {error}")

for name, url in PAGES.items():
    try:
        response = get(url)
        (RAW / f"{name}.html").write_bytes(response.content)
        soup = BeautifulSoup(response.text, "html.parser")
        links = [
            (" ".join(anchor.get_text(" ", strip=True).split()), urljoin(response.url, anchor["href"]))
            for anchor in soup.find_all("a", href=True)
        ]
        REPORT.append(f"LINKS {name}: {len(links)}")
        for text, href in links:
            key = f"{text} {href}".lower()
            wanted = (
                (name == "fsa_food_catalogue" and ("abridged" in key or "food and you" in key))
                or (name == "nhs_rtt" and "overview timeseries" in key)
                or (name == "ons_housing_catalogue" and href.lower().endswith((".xlsx", ".xls")))
                or (name == "ons_trust_catalogue" and href.lower().endswith((".xlsx", ".xls")))
                or (name == "ons_wellbeing_catalogue" and href.lower().endswith((".xlsx", ".xls")))
                or (name == "nrw_portal" and ("warning" in key or "flood" in key))
            )
            if wanted:
                REPORT.append(f"  CANDIDATE text={text!r} href={href}")
        if name == "nhs_rtt":
            priority = [(text, href) for text, href in links if "overview timeseries" in f"{text} {href}".lower()]
            if priority:
                text, href = priority[0]
                file_response = get(href)
                path = safe_name(name, file_response.url, file_response.headers.get("content-type", ""))
                path.write_bytes(file_response.content)
                selected[name] = {
                    "url": file_response.url,
                    "link_text": text,
                    "path": str(path),
                    "content_type": file_response.headers.get("content-type"),
                    "bytes": len(file_response.content),
                }
    except Exception as error:
        REPORT.append(f"ERROR {name}: {type(error).__name__}: {error}")

(OUT / "selected.json").write_text(json.dumps(selected, indent=2) + "\n")

for name, metadata in selected.items():
    path = Path(metadata["path"])
    preview = [f"## {name}", json.dumps(metadata, indent=2)]
    try:
        if path.suffix.lower() in (".csv", ".json", ".html", ".txt"):
            text = path.read_text(errors="replace")
            lines = text.splitlines()
            preview.append(f"lines={len(lines)}")
            preview.extend(lines[:50])
            for pattern in (
                "United Kingdom", "food security", "low food", "very low", "incomplete",
                "England", "2024-25", "2025", "trust", "Wave 11"
            ):
                matches = [line for line in lines if pattern.lower() in line.lower()][:20]
                if matches:
                    preview.append(f"-- matches: {pattern}")
                    preview.extend(matches)
        elif path.suffix.lower() == ".xlsx":
            import openpyxl
            book = openpyxl.load_workbook(path, read_only=True, data_only=True)
            preview.append("sheets=" + json.dumps(book.sheetnames))
            for sheet in book.worksheets:
                preview.append(f"-- sheet {sheet.title!r} rows={sheet.max_row} cols={sheet.max_column}")
                for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 40), values_only=True):
                    preview.append("\t".join("" if value is None else str(value) for value in row[:24]))
        elif path.suffix.lower() == ".xls":
            import xlrd
            book = xlrd.open_workbook(path)
            preview.append("sheets=" + json.dumps(book.sheet_names()))
            for sheet in book.sheets():
                preview.append(f"-- sheet {sheet.name!r} rows={sheet.nrows} cols={sheet.ncols}")
                for index in range(min(sheet.nrows, 40)):
                    preview.append("\t".join(str(value) for value in sheet.row_values(index)[:24]))
    except Exception as error:
        preview.append(f"INSPECTION ERROR {type(error).__name__}: {error}")
    (OUT / f"{name}.txt").write_text("\n".join(preview) + "\n")

(OUT / "report.txt").write_text("\n".join(REPORT) + "\n")
